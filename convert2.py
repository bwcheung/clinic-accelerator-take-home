import json
from openpyxl import load_workbook

result = {}

wb = load_workbook("Scoreboard Test.xlsx", data_only=True)
ws = wb.active

def format_value(cell):
  value = cell.value
  fmt = cell.number_format or ""

  if value is None:
    return None

  if "$" in fmt and isinstance(value, (int, float)):
    return f"${value:,.2f}"

  if "%" in fmt and isinstance(value, (int, float)):
    return f"{round(value * 100)}%"

  return value

rows = list(ws.iter_rows())
rows = [r for idx, r in enumerate(rows) if idx not in [0, 5, 6]]

headers = rows[:4]
data = rows[4:]

valid_cols = []
for i, cell in enumerate(headers[0]):
  val = cell.value
  if val is not None and str(val).strip() != "" and i != 0:
    valid_cols.append(i)

for row in data:
  if not row or row[0].value is None:
    continue

  date_cell = row[0].value

  if hasattr(date_cell, "strftime"):
    date = date_cell.strftime("%B %d, %Y")
  else:
    date = str(date_cell)

  result[date] = []

  for i in valid_cols:
    cell = row[i] if i < len(row) else None

    value = format_value(cell) if cell else None

    col_name = headers[0][i].value
    focus = headers[1][i].value if i < len(headers[1]) else None
    source = headers[2][i].value if i < len(headers[2]) else None
    role = headers[3][i].value if i < len(headers[3]) else None

    result[date].append({
      str(col_name): {
        "value": value,
        "focus": focus,
        "source": source,
        "Role": role
      }
    })

with open("output.json", "w") as file:
  json.dump(result, file, indent=2, default=str)